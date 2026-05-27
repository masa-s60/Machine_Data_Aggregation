from pathlib import Path
from PIL import Image
from openpyxl import load_workbook
from datetime import datetime
import pytesseract


# ---スクショデータ取得 and リネーム---

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

screenshot_folder = Path(r"C:\Users\fsmn2\OneDrive\画像\スクリーンショット")

screenshot_files = list(screenshot_folder.glob("*.png"))

latest_screenshot_file = max(screenshot_files, key=lambda screenshot_file: screenshot_file.stat().st_mtime)

target_date = input("日付を入力 YYYYMMDD: ")

screenshot_file_name = f"{target_date}_rotation.png"

screenshot_file_path = (latest_screenshot_file.parent / screenshot_file_name)

latest_screenshot_file.rename(screenshot_file_path)

print(screenshot_file_path)


# ---スクショから台データの取得---

img = Image.open(screenshot_file_path)

total_rotation_rect = img.crop((650, 565, 715, 835))
max_payout = img.crop((830, 565, 900, 835))

total_rotation_rect.show()
max_payout.show()

total_rotation_list = pytesseract.image_to_string(total_rotation_rect)
max_payout_list = pytesseract.image_to_string(max_payout)

print(total_rotation_list)
print(max_payout_list)

# ---Excelデータ取得、対象セル探索---

target_date = screenshot_file_name.split("_")[0]

target_date_obj = datetime.strptime(target_date, "%Y%m%d").date()

excel_date_text = f"{target_date[:4]}/{int(target_date[4:6])}/{int(target_date[6:])}"

print(excel_date_text)

excel_path = r"C:\Users\fsmn2\OneDrive\ドキュメント\Aggregated_Data\machine_data_aggregation.xlsx"

workbook = load_workbook(excel_path)

worksheet = workbook["マイジャグ"]

target_row = None

for row in range(1, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=1).value

    if hasattr(cell_value, "date"):
        cell_date = cell_value.date()
    else:
        continue

    if cell_date == target_date_obj:
        target_row = row
        break

print(f"対象行: {target_row}")


# ---OCR結果をリスト化---

total_rotation_values = [
    int(value) for value in total_rotation_list.split()
    if value.isdigit()
]

max_payout_values = [
    int(value) for value in max_payout_list.split()
    if value.isdigit()
]

print(total_rotation_values)
print(max_payout_values)


# ---Excelへ書き込み---

rotation_start_column = 4
max_payout_start_column = 11

for index, rotation_value in enumerate(total_rotation_values):

    target_column = rotation_start_column + index

    worksheet.cell(
        row=target_row,
        column=target_column
    ).value = rotation_value


for index, payout_value in enumerate(max_payout_values):

    target_column = max_payout_start_column + index

    worksheet.cell(
        row=target_row,
        column=target_column
    ).value = payout_value

workbook.save(excel_path)

print("書き込み完了")

import os
os.startfile(excel_path)